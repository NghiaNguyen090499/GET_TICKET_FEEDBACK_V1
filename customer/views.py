from django.shortcuts import render, redirect, get_object_or_404
from customer.models import *
from django.http import JsonResponse, HttpResponse
from collections import defaultdict
from django.urls import reverse
import time
from django.db.models import Max
from datetime import datetime
import qrcode
import qrcode.image.svg
from io import BytesIO
from django.conf import settings
from qrcode import make
from QRCODE.settings import MEDIA_ROOT
from django.views.decorators.csrf import csrf_exempt
from mailmerge import MailMerge
from docx import Document
from openpyxl import load_workbook
from num2words import num2words
from django.contrib.auth import authenticate, login
import subprocess
import win32com.client as win32 
import pythoncom  # Thêm import pythoncom
from openpyxl import Workbook
from django.db import connection
from openpyxl import Workbook
from gtts import gTTS
import tempfile
import os
import pygame
from license.license_manager import LicenseManager
from django.utils import timezone
from collections import defaultdict
from django.db.models import Count
license_manager = LicenseManager()


# Đây là một ví dụ về một hàm cần một đối số, nhưng khi gọi lại nó, không truyền đối số.
def test(request):
    if license_manager.is_valid_license():
        license_info = license_manager.get_license_info()
        print(f"License is valid. Expiration date: {license_info.expiration_date}")
    else:
        print("Invalid license or license has expired.")
    
    return HttpResponse('a')

def trang_chu_2(request):

    nhan_vien = Employee.objects.all()
    return render(request,'Qrcodes/customer/trang_chu_lay_so.html',{'nhan_vien':nhan_vien})
    
def dang_nhap(request):
    result_login=''
    if request.POST.get('btnDangNhap'):
    # Gán biến
        user_name = request.POST.get('user_name')
        mat_khau = request.POST.get('mat_khau')

        # Xử lý đọc thông tin từ CSDL
        nguoi_dung = Customer.objects.filter(user_name=user_name, password=mat_khau)
        if nguoi_dung.count() > 0:
            dict_nguoi_dung = nguoi_dung.values()[0]

            del(dict_nguoi_dung['password'])
            request.session['s_khachhang'] = dict_nguoi_dung
            return redirect('customer:trang-chu')
        else:
            result_login = '''
            <div class="alert alert-danger" role="alert">
                Đăng nhập thất bại. Vui lòng kiểm lại thông tin
            </div>
            '''

    return render(request, 'Qrcodes/login.html', {
    
    'result_login': result_login,
   
    })

def chon_dich_vu(request):
    dich_vu= Service.objects.all()
    nhan_vien=Employee.objects.all()
    services=Service.objects.all() 
    thoi_gian= Manage.objects.get(id=1)
    for service in services:
        context = {'initial_service_status': service.is_active}

 
    return render(request,'Qrcodes/customer/index_1.html',{'dich_vu':dich_vu, 'context':context,'nhan_vien':nhan_vien,'thoi_gian':thoi_gian})

def danh_gia(request,pk,ticket_number):
    nhan_vien = Employee.objects.get(assigned_service_id=pk)
    khach_hang = KhachHang.objects.filter(name__isnull=False,is_called=False, employee_id=pk, ticket_number=ticket_number ).first()
    url = reverse('customer:nhan-vien', args=[nhan_vien.assigned_service_id])
    danh_gia = CapNhatDuLieu.objects.all()
    hang_can_xoa = DanhSachCho.objects.get(ticket_number=ticket_number)
    print(hang_can_xoa)
# Xóa hàng

    if khach_hang:
        employee = Service.objects.get(pk=pk)
        if request.method == 'POST':   
            danh_gia = request.POST.get('feedback')
            # first_customer = KhachHang.objects.filter(is_called=False,is_calling=True, employee_id=pk).first()
            first_customer = KhachHang.objects.get(ticket_number=ticket_number)
            if first_customer:
                first_customer.is_calling = True  # Đặt is_called thành True để đánh dấu đã gọi
                first_customer.is_called = True # Đặt is_called thành True để đánh dấu đã gọi
                first_customer.feedback = danh_gia # Đặt is_called thành True để đánh dấu đã gọi
                first_customer.save()
                thong_ke = ThongKe.objects.create(ticket_number=ticket_number)
                thong_ke.name = first_customer.name
                thong_ke.dia_chi=first_customer.dia_chi
                thong_ke.ticket_number=first_customer.ticket_number
                thong_ke.feedback=first_customer.feedback
                thong_ke.service=first_customer.service
                thong_ke.ngay_sinh = first_customer.ngay_sinh
                thong_ke.so_cccd = first_customer.so_cccd
                
                thoi_gian_3 = timezone.now()
                formatted_time = thoi_gian_3.strftime("%d/%m/%Y %H:%M")  # Thêm giờ và phút vào định dạng chuỗi
                ngay_thang_dt = datetime.strptime(formatted_time, "%d/%m/%Y %H:%M")

                # Gán giá trị vào trường DateTimeField
                thong_ke.ngay_quet = ngay_thang_dt
                thong_ke.save()
                try:
                    hang_can_xoa = DanhSachCho.objects.get(ticket_number=ticket_number)
                    hang_can_xoa.delete()
                    
                except:
                    pass
                return redirect(url)
            else:
            # Nếu không có hàng nào có is_called=False, tạo một hàng mới
                return redirect(url)
            
    else:
        return redirect(url)
    return render(request,'Qrcodes/customer/danh_gia_html',{'khach_hang':khach_hang,'danh_gia':danh_gia})



         
def get_ticket(request, pk):
    context = {}
    img_name={}
    dich_vu = Service.objects.get(pk=pk)
    data_json_1=''
    ho_va_ten=''
    so_thu_tu='' 
    dia_chi=''
    sothutu=KhachHang()
    thoi_gian_1= Manage.objects.get(id=1)
    thoi_gian_1=thoi_gian_1.time
    
    if request.method == 'POST' and 'scan_data_1' in request.POST:
        sothutu = KhachHang.objects.all()
        thoi_gian_1= Manage.objects.get(id=1)
        thoi_gian_1=thoi_gian_1.time
        sothutu = KhachHang.objects.filter(is_calling=False).first()
        thong_tin=request.POST.get('thong_tin')
        try :
            thong_tin = thong_tin.rstrip()
            thong_tin=thong_tin.replace('|',"*")
            thong_tin=thong_tin.split('*')
            so_cccd = thong_tin[0]
            ho_va_ten = thong_tin[2].upper()
            ngay_sinh = thong_tin[3]
            ngay_sinh= datetime.strptime(ngay_sinh, "%d%m%Y")
            ngay_sinh = datetime.strftime(ngay_sinh, "%d/%m/%Y")
            gioi_tinh = thong_tin[4]
            dia_chi = thong_tin[5]
            ngay_cap = thong_tin[6]
            ngay_cap= datetime.strptime(ngay_cap, "%d%m%Y")
            ngay_cap = datetime.strftime(ngay_cap, "%d/%m/%Y")
            thanh_pho = dia_chi.replace('\x00','').split(',')
            thanh_pho_chuan=thanh_pho[-1]
            thoi_gian = datetime.now()
            thoi_gian= thoi_gian.strftime("%d/%m/%Y %H:%M")
            
            so_thu_tu = sothutu.ticket_number
            data_json_1 = {
            'ho_va_ten':ho_va_ten,
            'so_cccd': so_cccd,
            'dia_chi':dia_chi,
            'so_thu_tu': so_thu_tu,
            'ngay_sinh': ngay_sinh,
            'dich_vu': dich_vu.name,
            'gioi_tinh':gioi_tinh.replace('\x00',''),
            'dia_chi': dia_chi.replace('\x00',''),
            'ngay_cap':ngay_cap.replace('\x00',''),
            'thanh_pho': thanh_pho_chuan.replace('\x00','')
            
        }
        except:
            thong_bao = '''
                <div class="alert alert-success text-center display-3" role="alert">
                    Vui lòng chọn đúng mã QR trên căn cước công dân hoặc VneID
                </div>
                '''
            return render(request, 'Qrcodes/customer/lay_so.html',{'thong_bao':thong_bao,'thoi_gian':thoi_gian_1})      
        first_customer = KhachHang.objects.filter(is_calling=False).first()
        employee = Service.objects.get(pk=pk)
        if first_customer:
    # Cập nhật giá trị 'name' cho hàng đầu tiên
            first_customer.name = ho_va_ten
            first_customer.dia_chi = dia_chi
            first_customer.is_calling = True  # Đặt is_called thành True để đánh dấu đã gọi
            first_customer.employee_id = employee  # Đặt is_called thành True để đánh dấu đã gọi
            first_customer.service = employee.name
            first_customer.ngay_sinh = ngay_sinh
            first_customer.so_cccd = so_cccd# Đặt is_called thành True để đánh dấu đã gọi
            first_customer.save()
        else:
 
            pass
        request.session['scan_data_1'] = data_json_1
        context = {}
        url = reverse('customer:gui-khach-hang', args=[pk,sothutu.ticket_number])
        data = f"fc71-115-78-6-130.ngrok-free.app/{url}"
        img = make(data)
        img_name = f'{so_thu_tu}.png'
        img.save(settings.MEDIA_ROOT + '/' + img_name)
        context = {
            'img_name': img
        }
        data_1 = request.session.get('scan_data_1')
        if data_1 : 
            thoi_gian = datetime.now()
            thoi_gian= thoi_gian.strftime("%d/%m/%Y %H:%M")
            file_path = MEDIA_ROOT + 'QRcodes/reports/data/data_qr.xlsx'
            
            # Mở file Excel hiện có
            workbook = load_workbook(filename=file_path)
            # Chọn sheet bạn muốn thêm dữ liệu (ví dụ: 'Sheet1')
            sheet = workbook['Sheet1']
            # Dữ liệu mới bạn muốn thêm
            new_data = [thoi_gian ,
                        data_1["so_cccd"],       
                        data_1["ho_va_ten"],
                        data_1["ngay_sinh"],
                        data_1["gioi_tinh"],
                        data_1["dia_chi"],
                        data_1["ngay_cap"],
                    ]    
            # Thêm dữ liệu vào hàng mới trong sheet
            sheet.append(new_data)
         # Lưu lại file Excel sau khi thêm dữ liệu
            workbook.save(filename=file_path)
         
    return render(request, 'Qrcodes/customer/lay_so.html',{'so_thu_tu': sothutu.ticket_number,
                                                           'data_1': data_json_1,
                                                           'ho_va_ten':ho_va_ten,
                                                           'dia_chi':dia_chi,
                                                           'dich_vu':dich_vu.name,
                                                            'so_thu_tu': so_thu_tu,
                                                            'context':context,
                                                            'img_name':img_name,
                                                            'thoi_gian':thoi_gian_1}
                  )


def nhan_vien(request,assigned_service_id):
    employee = Employee.objects.get(pk=assigned_service_id)
    assigned_service_name = employee.assigned_service.name
    default_trangthai = TrangThai.objects.get(pk=1)
    khachhang = KhachHang.objects.filter(pk=assigned_service_id)
    tong_khach_hang = KhachHang.objects.filter(is_called=False)
    service_and_order = defaultdict(list)
    danh_sach_cho = DanhSachCho.objects.filter(employee=assigned_service_id)
    print(len(danh_sach_cho))
    print(danh_sach_cho)
    for khach_hang in tong_khach_hang:
        if khach_hang.service:
            service_and_order[khach_hang.service].append(khach_hang.ticket_number)
    for service, orders in service_and_order.items():
        print(f"Dịch vụ: {service} - Số thứ tự đầu tiên: {orders[0]}")

    for khachhang in khachhang:
        if khachhang.trang_thai is None:
            KhachHang.objects.filter(trang_thai__isnull=True).update(trang_thai=default_trangthai)
    
    waiting_customers = KhachHang.objects.filter(name__isnull=False, employee_id=assigned_service_id ,is_called=False,trang_thai=True)
    
    service=Service.objects.get(pk=assigned_service_id)
    trang_thai= ''
    
    if employee.is_active ==True:    
        trang_thai = 'Hoạt động'
    elif employee.is_active ==False:  
        trang_thai = ' Không Hoạt động'
  
    if request.method == 'POST' and 'scan_data_1' in request.POST: 
        employee=Employee.objects.get(pk=assigned_service_id)
        service=Service.objects.get(pk=assigned_service_id)
        if employee.is_active ==True:
            employee.is_active = False
            service.is_active = employee.is_active
            trang_thai = 'Không hoạt động'
            employee.save()
            service.save()
        elif employee.is_active ==False:
            employee.is_active = True
            service.is_active = employee.is_active
            trang_thai = 'Hoạt động'
            employee.save()
            service.save()

    count = 0
    
    for i in waiting_customers:
        count +=1
        

    return render(request,'Qrcodes/customer/nhan_vien_hanh_chinh.html',{'loai_dich_vu': assigned_service_name,
                                                                        'ten_nhan_vien' :employee.name,
                                                                        'khach_hang_cho':waiting_customers,
                                                                        'employee':employee,
                                                                          'dem': count,
                                                                          'trang_thai':trang_thai,
                                                                          'quay':employee.pk,
                                                                          'danh_sach_cho':danh_sach_cho,
                                                                          
                                                                          'tong_khach_hang':tong_khach_hang,
                                                                          'service':service_and_order.items(),
                                                                          }
                                                                        )



def bang_khach_hang(request):
    dich_vu= Service.objects.all
    khach_hang_list = KhachHang.objects.filter(name__isnull=False,is_called=False)
    grouped_khach_hang = defaultdict(list)
    for khach_hang in khach_hang_list:
        # Group khach_hang theo dịch vụ
        grouped_khach_hang[khach_hang.service].append(khach_hang)
    
    # Chuyển từ điển thành danh sách tuples để truyền vào template
    grouped_khach_hang_list = list(grouped_khach_hang.items())
    

    
    return render(request,'Qrcodes/customer/bang_khach_hang.html',{
                  'grouped_khach_hang_list': grouped_khach_hang_list,
                      
                         'dich_vu':dich_vu}
                        )

    
    
    
def gui_khach_hang(request,pk,ticket_number):
    
    so_thu_tu = KhachHang.objects.filter(name__isnull=False, is_called=False,employee_id=pk, feedback=None).order_by('ticket_number')
    ho_va_ten=KhachHang.objects.filter(ticket_number=ticket_number,employee_id=pk,feedback=None).first()
    first_ticket_number = so_thu_tu.first().ticket_number
    danh_sach_so_thu_tu = KhachHang.objects.filter(is_called=False, employee_id=pk ).values_list('ticket_number', flat=True)
    danh_sach_so_thu_tu = [int(stt) for stt in danh_sach_so_thu_tu]
    # Đếm số lượng số thứ tự liền trước
    so_luong_lien_truoc = len([stt for stt in danh_sach_so_thu_tu if stt < ticket_number])


    if ticket_number == first_ticket_number:
        dem = 0
    else:
        if ho_va_ten:  
            dem = so_luong_lien_truoc
            if dem < 0:
                dem = 'Đã xử lý xong'
        else:
            dem=''
    khach_hang = KhachHang.objects.filter(name__isnull=False,is_called=False, employee_id=pk, ticket_number=ticket_number)
    thoi_gian = datetime.now()
    time = thoi_gian.strftime("%H:%M:%S")
    
    
    return render(request,'Qrcodes/customer/gui_khach_hang.html',{
                            'khach_hang': khach_hang,
                            'thoi_gian' :time,
                            'dem': dem,

                         
                         } )
    
        
def generate_qr_code(request,pk,ticket_number):
    # Tạo đường dẫn URL với các tham số
    context = {}
    if request.method == "POST":
        data = request.POST['qr_text']
        img = make(data)
        test='a'
        img_name = f'{test}.png'
        img.save(settings.MEDIA_ROOT + '/' + img_name)
        context = {
            'img_name': img_name
        }
        
       

    return render(request, "Qrcodes/customer/qr.html",context)

def tong_hop(request):  

    if 's_khachhang' not in request.session:
        return redirect('customer:dang-nhap')
    
    selected_columns = ['name', 'ticket_number', 'service','feedback']
    # Thực hiện truy vấn SQL để lấy dữ liệu từ bảng của bạn với các cột đã chọn
    with connection.cursor() as cursor:
        cursor.execute(f"SELECT {', '.join(selected_columns)} FROM customer_khachhang")
        columns = [col[0] for col in cursor.description]
        data = cursor.fetchall()

    # Tạo Workbook và WorkSheet
    wb = Workbook()
    ws = wb.active

    # Thêm tên cột vào WorkSheet
    ws.append(selected_columns)

    # Thêm dữ liệu vào WorkSheet
    for row in data:
        ws.append(row)

    # Lưu Workbook thành tệp Excel
    current_date = datetime.now().strftime('%Y-%m-%d')
    file_name= f'output_{current_date}.xlsx'
    folder_report = MEDIA_ROOT + 'QRcodes\\reports\\'
    path_report = folder_report + file_name
    wb.save(path_report)
    
    if 's_khachhang' not in request.session:
        return redirect('customer:dang-nhap')
    
    KhachHang.objects.all().delete()
  
    # Xóa tất cả bản ghi có trường da_phuc_vu=True
    KhachHang.objects.update(is_called=False)
    
    
    for i in range(1,201):
 
        ticket_number = KhachHang.objects.create(ticket_number=i, is_called=False, is_calling=False)
        
    del request.session['s_khachhang']
    
    return redirect('customer:trang-chu')
    



def text_to_speech(request):
    audio_data=''
    if request.method == 'POST':
        text_to_read = request.POST.get('text', '')
        # Sử dụng gTTS để tạo dữ liệu audio từ văn bản
        tts = gTTS(text_to_read, lang='vi', slow=False)
        
        # Tạo một tệp tạm thời để lưu trữ dữ liệu audio
        temp_file_path = tempfile.mktemp(suffix='.mp3')
        tts.save(temp_file_path)
        print(temp_file_path)
        
        pygame.init()


        pygame.mixer.music.load(temp_file_path)
        pygame.mixer.music.play()

# Giữ chương trình chạy cho đến khi nhạc kết thúc
        pygame.event.wait()
    

    return HttpResponse(audio_data, content_type='audio/mpeg')


def toggle_employee_status(request, employee_id):
    employee = get_object_or_404(Employee, pk=employee_id)
    service = get_object_or_404(Service, pk=employee_id)
    service.is_active=employee.is_active
    service.save()
    # Toggle trạng thái
  
    response_data = {
        'employee_status': employee.is_active,
        'service_status': service.is_active,
    }
    
    return JsonResponse(response_data)




def dem(request):
    danh_sach = ThongKe.objects.all()
    du_lieu_feedback = CapNhatDuLieu.objects.all()

    # Lặp qua từng đối tượng trong du_lieu_feedback và đếm số lượng phản hồi cho mỗi id
    for du_lieu in du_lieu_feedback:
        id = du_lieu.id
        count = 0
        for feedback in danh_sach:
            if feedback.feedback == du_lieu.feedback:
                count += 1

        # Kiểm tra xem đã tồn tại bản ghi cho id trong FeedbackCount hay chưa
        try:
            feedback_count = CapNhatDuLieu.objects.get(id=id)
            feedback_count.tong = count
            feedback_count.save()  # Cập nhật bản ghi đã tồn tại
        except CapNhatDuLieu.DoesNotExist:
            CapNhatDuLieu.objects.create(id=id, count=count)  # Tạo mới bản ghi nếu chưa tồn tại
    # Gửi số lượng phản hồi cho mỗi id đến giao diện
 





def feedback_chart_data(request):
    # Lấy dữ liệu feedback theo ngày tháng, sắp xếp theo ngày tháng
    feedback_data = ThongKe.objects.filter(ngay_quet__isnull=False).values('ngay_quet__date', 'feedback').annotate(count=Count('id')).order_by('ngay_quet__date')

    # Tạo danh sách để lưu thông tin feedback cho mỗi ngày
    data = []
    for item in feedback_data:
        date_obj = item['ngay_quet__date']
        # Định dạng lại ngày tháng
        formatted_date = date_obj.strftime('%d-%b-%Y')

        # Tạo một đối tượng mới để lưu thông tin feedback
        feedback_item = {
            'date': formatted_date,
            'count': item['count'],
            'feedbacks': {item['feedback']: item['count']}
        }

        # Kiểm tra xem ngày đã tồn tại trong danh sách hay chưa
        existing_item = next((x for x in data if x['date'] == formatted_date), None)
        if existing_item:
            existing_item['count'] += item['count']
            if item['feedback'] not in existing_item['feedbacks']:
                existing_item['feedbacks'][item['feedback']] = item['count']
            else:
                existing_item['feedbacks'][item['feedback']] += item['count']
        else:
            data.append(feedback_item)

    return JsonResponse({'data': data})


def move_to_waiting_list(request, khachhang_id):
    khachhang = get_object_or_404(KhachHang, ticket_number=khachhang_id)
   
    if request.method == "POST":
        khachhang = get_object_or_404(KhachHang, ticket_number=khachhang_id)
        nhan_vien=khachhang.employee_id
        url = reverse('customer:nhan-vien', args=[khachhang.employee_id])
        if khachhang:
            khachhang.trang_thai = False
            khachhang.save()
            DanhSachCho.objects.create(khach_hang=khachhang,name=khachhang.name,ticket_number=khachhang_id,employee=khachhang.employee_id)
            redirect(url)
        else:
            return JsonResponse({'error': 'Khach hang khong dang duoc goi.'}, status=400)
    
    return  redirect(url)
    
    